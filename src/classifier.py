import json
import os
from werkzeug.datastructures import FileStorage
from pdfminer.high_level import extract_text
from docx import Document
import openpyxl
from PIL import Image
from rapidfuzz import fuzz
import pytesseract
import magic
import io
import pickle

pytesseract.pytesseract.tesseract_cmd = r'/usr/bin/tesseract'

# Load ML model and vectorizer
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH = os.path.join(BASE_DIR, "models", "model.pkl")
VECTORIZER_PATH = os.path.join(BASE_DIR, "models", "vectorizer.pkl")
CATEGORY_FILE = os.path.join(BASE_DIR, "models", "categories.json")

with open(MODEL_PATH, "rb") as f:
    model = pickle.load(f)
with open(VECTORIZER_PATH, "rb") as f:
    vectorizer = pickle.load(f)

def load_categories():
    """Load categories and their keywords from categories.json."""
    with open(CATEGORY_FILE, "r") as f:
        categories = json.load(f)
    return categories

def save_category(name, keywords):
    """Save a new category to the categories.json file."""
    conflicts = fuzzy_match_category(name, keywords)
    if conflicts:
        return conflicts

    categories = load_categories()
    categories[name] = keywords
    with open(CATEGORY_FILE, "w") as f:
        json.dump(categories, f, indent=4)
    return None

def extract_text_from_pdf(file_bytes):
    with open("temp.pdf", "wb") as f:
        f.write(file_bytes)
    return extract_text("temp.pdf")

def extract_text_from_image(file_bytes):
    image = Image.open(io.BytesIO(file_bytes))
    return pytesseract.image_to_string(image)

def extract_text_from_docx(file_bytes):
    doc = Document(io.BytesIO(file_bytes))
    text = ""
    for para in doc.paragraphs:
        text += para.text + "\n"
    return text

def extract_text_from_xlsx(file_bytes):
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text

def fuzzy_match(text, keyword, threshold=80):
    score = fuzz.token_set_ratio(text, keyword)
    print(score)
    return score >= threshold

def fuzzy_match_category(new_category_name, new_keywords, threshold = 80):
    categories = load_categories()
    conflicts = []

    #first, lets check for similarities in category names

    for current_cat, current_keywords in categories.items(): 
        if fuzzy_match(new_category_name, current_cat, threshold=70):
            conflicts.append( f"Category name '{new_category_name}' is too similar to existing category '{current_cat}'")

    #next, lets check for similarities in keywords 
    for new_keyword in new_keywords:
        for current_cat, current_keywords in categories.items():
            for current_keyword in current_keywords:
                if fuzzy_match(new_keyword, current_keyword, threshold=80):
                    conflicts.append(f"Keyword '{new_keyword}' is too similar to '{current_keyword}' in category '{current_cat}'")
    
    return conflicts


def classify_text(text):
    text = text.lower()

    categories = load_categories()

    # Iterate over categories and their keywords
    for category, keywords in categories.items():
        for keyword in keywords:
            if any(fuzzy_match(line, keyword) for line in text.splitlines()):
                return category

    return "unknown file"

def ml_classify_text(text):
    X = vectorizer.transform([text])
    prediction = model.predict(X)
    return prediction[0]

def classify_file(file: FileStorage):
    filename = file.filename.lower()
    file_bytes = file.read()
    mime_type = magic.from_buffer(file_bytes, mime=True)

    if mime_type == "application/pdf":
        text = extract_text_from_pdf(file_bytes)
    elif mime_type.startswith("image/"):
        text = extract_text_from_image(file_bytes)
    elif filename.endswith(".docx"):
        text = extract_text_from_docx(file_bytes)
    elif filename.endswith(".xlsx"):
        text = extract_text_from_xlsx(file_bytes)
    elif mime_type == "text/plain":
        text = file_bytes.decode("utf-8")
    else:
        return "unsupported format"

    label = classify_text(text)
    if label == "unknown file":
        label = ml_classify_text(text)
    return label
